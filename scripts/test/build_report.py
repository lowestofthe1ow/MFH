"""
Builds an Excel report from author-adaptation ablation results.

Expects a folder layout like:

    finetune_checkpoints/
      author_0/
        digital/ablation_results.json
        phys_norush/ablation_results.json
      author_1/
        digital/ablation_results.json
        phys_norush/ablation_results.json
        phys_rush/ablation_results.json

(each ablation_results.json produced by the finetune_custom_author.py-style
driver: {"baseline": {"exprate":.., "le1":.., "le2":.., "cer":..},
"ablation": {"5": {"961": {"exprate":.., "le1":.., "le2":.., "cer":..}, ...},
"8": {...}, ...}, ...})

Produces one .xlsx with two sheets:
  - "Summary":   the human-facing report, laid out to match the target
                 mockup (one header block per author, one merged-label
                 block per handwriting category, rows = base + each
                 finetune size). Reports ExpRate, "<=1 error", and
                 "<=2 error" (all higher-is-better, matching the MFH paper's
                 tolerance metrics) plus CER (lower-is-better).
  - "Raw Data":  every per-seed metric value. The Summary sheet's metric/
                 std cells are live formulas (AVERAGE/STDEV) referencing
                 this sheet, and the gain/improvement cells are formulas
                 referencing other Summary cells -- nothing in the numeric
                 columns is a hardcoded Python literal.

Usage:
    python build_report.py --root finetune_checkpoints --out ablation_report.xlsx
"""

import json
import argparse
from collections import namedtuple
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Category discovery / labeling
# ---------------------------------------------------------------------------
CATEGORY_ORDER = ["digital", "phys_norush", "phys_rush"]
CATEGORY_LABELS = {
    "digital": "Handwritten (Digital)",
    "phys_norush": "Handwritten (Physical - Not rushed)",
    "phys_rush": "Handwritten (Physical - Rushed)",
}

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10)
BASE_FONT = Font(name=FONT_NAME, size=10)
LABEL_FONT = Font(name=FONT_NAME, size=10)
DELTA_FONT = Font(name=FONT_NAME, size=10, color="ED7D31")  # orange, matches mockup
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FMT = "0.00%"

# ---------------------------------------------------------------------------
# Metric table -- the single source of truth for which metrics appear, in
# what order, and whether higher or lower is better. Add a metric here (and
# to load_ablation_json's per-metric extraction) and the rest of the sheet
# layout follows automatically -- no other code needs to change.
#
#   key             : dict key in ablation_results.json
#   label           : display name in the Summary sheet
#   raw_col         : column index (1-based) for this metric in "Raw Data"
#   higher_better   : True -> gain = current - reference (ExpRate-style)
#                      False -> improvement = reference - current (CER-style)
# ---------------------------------------------------------------------------
_MetricDef = namedtuple("_MetricDef", "key label raw_col higher_better")
METRIC_DEFS = [
    _MetricDef("exprate", "ExpRate", 5, True),
    _MetricDef("le1", "<=1 Error", 6, True),
    _MetricDef("le2", "<=2 Error", 7, True),
    _MetricDef("cer", "CER", 8, False),
]

# Summary-sheet column layout, derived from METRIC_DEFS: each metric gets
# 4 columns (value, vs. base, vs. previous, std), laid out in METRIC_DEFS
# order after the two fixed leading columns (label, finetune size).
COL_LABEL = 1
COL_SIZE = 2
_Metric = namedtuple("_Metric", "key label higher_better value_col vs_base_col vs_prev_col std_col raw_col")
METRICS = []
for _idx, _m in enumerate(METRIC_DEFS):
    _base_col = COL_SIZE + 1 + _idx * 4
    METRICS.append(_Metric(
        key=_m.key, label=_m.label, higher_better=_m.higher_better,
        value_col=_base_col, vs_base_col=_base_col + 1, vs_prev_col=_base_col + 2,
        std_col=_base_col + 3, raw_col=_m.raw_col,
    ))
COLS = [COL_LABEL, COL_SIZE] + [c for m in METRICS for c in
                                 (m.value_col, m.vs_base_col, m.vs_prev_col, m.std_col)]


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def discover_results(root: Path):
    """
    Scans root/author_*/<category>/ablation_results.json and returns an
    ordered list of (author_dirname, category_dirname, json_path) tuples.
    Categories are ordered per CATEGORY_ORDER first, then any extras found
    on disk, alphabetically.
    """
    entries = []
    for author_dir in sorted((p for p in root.glob("author_*") if p.is_dir()), key=lambda p: p.name):
        present = {
            p.name for p in author_dir.iterdir()
            if p.is_dir() and (p / "ablation_results.json").exists()
        }
        ordered = [c for c in CATEGORY_ORDER if c in present]
        extra = sorted(present - set(ordered))
        for cat in ordered + extra:
            entries.append((author_dir.name, cat, author_dir / cat / "ablation_results.json"))
    return entries


def _require_metric(d, key, context):
    if key not in d:
        raise KeyError(
            f"'{key}' not found in {context}. This report script expects the "
            f"le1/le2 tolerance metrics added by the updated finetune_custom_author.py "
            f"-- rerun the ablation sweep to regenerate ablation_results.json with "
            f"these fields (older json files predate this metric and won't have it)."
        )
    return d[key]


def load_ablation_json(path: Path):
    with open(path) as f:
        data = json.load(f)

    baseline = data["baseline"]
    base = {m.key: _require_metric(baseline, m.key, f"{path} baseline") for m in METRICS}

    sizes = sorted(int(s) for s in data["ablation"].keys())
    per_size = {}
    for size in sizes:
        seed_entries = data["ablation"][str(size)]
        per_size[size] = {
            "seeds": list(seed_entries.keys()),
            **{
                m.key: [_require_metric(v, m.key, f"{path} size={size}") for v in seed_entries.values()]
                for m in METRICS
            },
        }
    return {"base": base, "sizes": sizes, "per_size": per_size}


# ---------------------------------------------------------------------------
# Sheet writing helpers
# ---------------------------------------------------------------------------

def _style(cell, font=BASE_FONT, fmt=None, align="right", fill=None):
    cell.font = font
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill


def write_header_block(ws, row, author_label):
    """
    Writes the 2-row header for one author's section, generalized over
    METRICS:
      row   : author_label | Finetune size | (blank) | "<Metric> mean gain/improvement" (merged over 3 cols) | ... (repeated per metric)
      row+1 : (blank)      | (blank)       | <Metric> | vs. base | vs. previous | std | ... (repeated per metric)
    """
    for m in METRICS:
        ws.merge_cells(start_row=row, start_column=m.vs_base_col, end_row=row, end_column=m.std_col)

    row1_values = {COL_LABEL: author_label, COL_SIZE: "Finetune size"}
    for m in METRICS:
        suffix = "mean gain" if m.higher_better else "mean improvement"
        row1_values[m.vs_base_col] = f"{m.label} {suffix}"
    for col in COLS:
        cell = ws.cell(row=row, column=col, value=row1_values.get(col))
        _style(cell, font=HEADER_FONT, align="left" if col == COL_LABEL else "center", fill=HEADER_FILL)

    row2_values = {}
    for m in METRICS:
        row2_values[m.value_col] = m.label
        row2_values[m.vs_base_col] = "vs. base"
        row2_values[m.vs_prev_col] = "vs. previous"
        row2_values[m.std_col] = "std"
    for col in COLS:
        cell = ws.cell(row=row + 1, column=col, value=row2_values.get(col))
        _style(cell, font=HEADER_FONT, align="center", fill=HEADER_FILL)

    return row + 2  # next writable row


def write_data_row(ws, row, size_label, values, base_row=None, prev_row=None, raw_ranges=None):
    """
    values:      dict metric_key -> formula string for the value cell
                 (e.g. "='Raw Data'!E12" for base, "=AVERAGE(...)" otherwise)
    base_row:    Summary-sheet row of the "base" row in this block, or None
                 if THIS row IS the base row (writes "-" for all deltas/std)
    prev_row:    Summary-sheet row of the previous ablation size (or the
                 base row, for the first ablation size)
    raw_ranges:  dict metric_key -> (start_raw_row, end_raw_row, n_seeds),
                 required whenever base_row is not None
    """
    _style(ws.cell(row=row, column=COL_SIZE, value=size_label), align="right")

    for m in METRICS:
        val_cell = ws.cell(row=row, column=m.value_col, value=values[m.key])
        _style(val_cell, fmt=PCT_FMT)

        if base_row is None:
            for col in (m.vs_base_col, m.vs_prev_col, m.std_col):
                cell = ws.cell(row=row, column=col, value="-")
                _style(cell, font=BASE_FONT, align="right")
            continue

        col_letter = get_column_letter(m.value_col)
        if m.higher_better:
            vs_base_formula = f"={col_letter}{row}-{col_letter}{base_row}"
            vs_prev_formula = f"={col_letter}{row}-{col_letter}{prev_row}"
        else:
            vs_base_formula = f"={col_letter}{base_row}-{col_letter}{row}"
            vs_prev_formula = f"={col_letter}{prev_row}-{col_letter}{row}"

        start_raw, end_raw, n_seeds = raw_ranges[m.key]
        raw_letter = get_column_letter(m.raw_col)
        std_formula = (
            f"=STDEV('Raw Data'!{raw_letter}{start_raw}:{raw_letter}{end_raw})"
            if n_seeds > 1 else None
        )

        for col, formula in ((m.vs_base_col, vs_base_formula), (m.vs_prev_col, vs_prev_formula),
                              (m.std_col, std_formula)):
            cell = ws.cell(row=row, column=col)
            if formula is None:
                cell.value = "-"
                _style(cell, font=BASE_FONT, align="right")
            else:
                cell.value = formula
                _style(cell, font=DELTA_FONT, fmt=PCT_FMT, align="right")


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------

def build_workbook(root: Path, out_path: Path):
    entries = discover_results(root)
    if not entries:
        raise FileNotFoundError(
            f"No ablation_results.json files found under "
            f"{root}/author_*/<category>/ablation_results.json"
        )

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    raw = wb.create_sheet("Raw Data")

    raw_headers = ["Author", "Category", "Finetune Size", "Seed"] + [m.label for m in METRICS]
    for c, h in enumerate(raw_headers, start=1):
        cell = raw.cell(row=1, column=c, value=h)
        _style(cell, font=HEADER_FONT, align="left", fill=HEADER_FILL)
    raw_row = 2

    summary_widths = {COL_LABEL: 30, COL_SIZE: 13}
    for m in METRICS:
        summary_widths[m.value_col] = 11
        summary_widths[m.vs_base_col] = 12
        summary_widths[m.vs_prev_col] = 13
        summary_widths[m.std_col] = 9
    for c, w in summary_widths.items():
        summary.column_dimensions[get_column_letter(c)].width = w

    srow = 1
    current_author = None

    for author, category, json_path in entries:
        data = load_ablation_json(json_path)

        if author != current_author:
            author_label = author.replace("_", " ").title()  # "author_0" -> "Author 0"
            srow = write_header_block(summary, srow, author_label)
            current_author = author

        block_start_row = srow
        base_row = srow

        # -- base row --
        raw.cell(row=raw_row, column=1, value=author)
        raw.cell(row=raw_row, column=2, value=CATEGORY_LABELS.get(category, category))
        raw.cell(row=raw_row, column=3, value="base")
        raw.cell(row=raw_row, column=4, value="-")
        for c in (1, 2, 3, 4):
            _style(raw.cell(row=raw_row, column=c), align="left")
        for m in METRICS:
            raw.cell(row=raw_row, column=m.raw_col, value=data["base"][m.key])
            _style(raw.cell(row=raw_row, column=m.raw_col), fmt=PCT_FMT)
        base_raw_row = raw_row
        raw_row += 1

        base_values = {m.key: f"='Raw Data'!{get_column_letter(m.raw_col)}{base_raw_row}" for m in METRICS}
        write_data_row(summary, base_row, size_label="base", values=base_values)
        srow += 1
        prev_row = base_row

        # -- ablation size rows (nested train sets, per finetune_custom_author.py) --
        for size in data["sizes"]:
            info = data["per_size"][size]
            start_raw = raw_row
            n_seeds = len(info["seeds"])
            for seed_idx, seed in enumerate(info["seeds"]):
                raw.cell(row=raw_row, column=1, value=author)
                raw.cell(row=raw_row, column=2, value=CATEGORY_LABELS.get(category, category))
                raw.cell(row=raw_row, column=3, value=size)
                raw.cell(row=raw_row, column=4, value=seed)
                for c in (1, 2, 3, 4):
                    _style(raw.cell(row=raw_row, column=c), align="left")
                for m in METRICS:
                    raw.cell(row=raw_row, column=m.raw_col, value=info[m.key][seed_idx])
                    _style(raw.cell(row=raw_row, column=m.raw_col), fmt=PCT_FMT)
                raw_row += 1
            end_raw = raw_row - 1

            this_row = srow
            values = {
                m.key: f"=AVERAGE('Raw Data'!{get_column_letter(m.raw_col)}{start_raw}:{get_column_letter(m.raw_col)}{end_raw})"
                for m in METRICS
            }
            raw_ranges = {m.key: (start_raw, end_raw, n_seeds) for m in METRICS}

            write_data_row(
                summary, this_row, size_label=size, values=values,
                base_row=base_row, prev_row=prev_row, raw_ranges=raw_ranges,
            )
            srow += 1
            prev_row = this_row

        block_end_row = srow - 1
        if block_end_row > block_start_row:
            summary.merge_cells(start_row=block_start_row, start_column=COL_LABEL,
                                 end_row=block_end_row, end_column=COL_LABEL)
        label_cell = summary.cell(row=block_start_row, column=COL_LABEL,
                                   value=CATEGORY_LABELS.get(category, category))
        _style(label_cell, font=LABEL_FONT, align="left")
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        srow += 1  # blank separator row between category blocks

    raw.column_dimensions["A"].width = 14
    raw.column_dimensions["B"].width = 32
    raw.column_dimensions["C"].width = 13
    raw.column_dimensions["D"].width = 8
    for m in METRICS:
        raw.column_dimensions[get_column_letter(m.raw_col)].width = 11

    wb.save(out_path)
    return out_path


def parse_args():
    p = argparse.ArgumentParser(description="Build an Excel ablation report from ablation_results.json files")
    p.add_argument("--root", default="finetune_checkpoints",
                    help="Folder containing author_*/<category>/ablation_results.json")
    p.add_argument("--out", default="ablation_report.xlsx", help="Output .xlsx path")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    out = build_workbook(Path(args.root), Path(args.out))
    print(f"Wrote {out}")
    print("Remember to run recalc.py on the output so cached formula values "
          "are populated (openpyxl writes formulas with no cached results).")